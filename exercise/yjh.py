#-*- coding:utf-8 -*-

import openpyxl

def main():
    dict = {}
    wb_end = openpyxl.Workbook()
    wb_end.create_sheet(u"总和")
    ws_end = wb_end.get_sheet_by_name(u"总和")
    wb_end.remove_sheet(wb_end.get_sheet_by_name("Sheet"))
    for x in range(1,110):
        try:
            sheet_name = ''
            wb = openpyxl.load_workbook('input/%s.xlsx'%str(x))
            sheet_names = wb.get_sheet_names()
            for name in sheet_names:
                if u"互联网攻击种类" in name:
                    sheet_name = name
                    break
            ws = wb.get_sheet_by_name(sheet_name)
            ws = list(ws)[1:]
            for row in ws:
                if row[0].value not in dict.keys():
                    dict[row[0].value] = int(row[1].value)
                else:
                    dict[row[0].value] += int(row[1].value)
        except:
            pass
    for key in dict.keys():
        ws_end.append([key, dict[key]])
    wb_end.save('ok.xlsx')
    wb_end.close()
    wb.close()

if __name__ == '__main__':
    main()