# -*- coding:utf-8 -*-

import openpyxl

class Excel_Creat(object):
    def __init__(self, time):
        self.time = time

    def excel_Filter(self):
        wb = openpyxl.Workbook()
        wb.save(u"outputFile/" + self.time + u"/数据筛选.xlsx")
        wb.close()

    def excel_Top5(self):
        wb = openpyxl.Workbook()
        wb.save(u"outputFile/" + self.time + u"/Top5.xlsx")

    def excel_Chart(self):
        wb = openpyxl.Workbook()
        wb.save(u"outputFile/" + self.time + u"/图表.xlsx")
        wb.close()

    def excel_Count(self):
        wb = openpyxl.Workbook()
        wb.save(u"outputFile/" + self.time + u"/统计.xlsx")
        wb.close()

    def run(self):
        self.excel_Filter()
        self.excel_Top5()
        self.excel_Count()
        self.excel_Chart()