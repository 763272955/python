# -*- coding:utf-8 -*-

import openpyxl
import _xlxs_csv
import datetime

class Data_Operat(object):
    def __init__(self):
        self.sys_all = None
        self.sys_other = None

    def create_Newsheet(self, filename, sheetname):
        wb = openpyxl.load_workbook(filename)
        for name in sheetname:
            wb.create_sheet(name)
        try:
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        except:
            pass
        return wb

    def get_Sheet(self, wb, name, title):
        ws = wb.get_sheet_by_name(name)
        ws.append(title)
        return ws

    def save_Excel(self, wb, filename):
        wb.save(filename)
        wb.close()

    def get_Data(self, rows):
        data = []
        for row in rows:
            data.append(row.value)
        return data

    def count_Dict(self, dict, key):
        if key not in dict.keys():
            dict[key] = 1
        else:
            dict[key] += 1
        return dict

    def get_Dict_data(self, ws, dict):
        data = []
        data_sort = []
        for key in dict.keys():
            data.append([key, dict[key]])
        for x in range(len(data)):
            if x == 0:
                data_sort.append(data[x])
                continue
            data_sort.append(data[x])
            for i in range(1, x+1):
                if int(data_sort[x+1-i][1]) > int(data_sort[x - i][1]):
                    a = data_sort[x-i]
                    data_sort[x-i] = data_sort[x+1-i]
                    data_sort[x+1-i] = a
        for x in range(len(data_sort)):
            ws.append(data_sort[x])
        return ws

    def system_Belong(self, ip, assetall):
        for all in assetall:
            if ip == all[0]:
                self.sys_all = all[1]
        return self.sys_all

    def data_Regular(self, response, record_type):
        test = 0
        if response != None:
            for type in record_type:
                result = type.findall(response)
                if len(result) != 0:
                    return False
        return True

    def rate_Count(self, time):
        time_ = []
        rate = {}
        wb_source = openpyxl.load_workbook("inputFile/" + time + "/IP_with_area.xlsx")
        wb_end = openpyxl.load_workbook("outputFile/" + time + u"/统计.xlsx")
        wb_end.create_sheet(u"IP(全)")
        ws_source = wb_source.get_sheet_by_name("Sheet")
        ws_end = wb_end.get_sheet_by_name(u"IP(全)")
        for i in range(2, 8):
            time_.append((datetime.datetime.now() - datetime.timedelta(days=i)).strftime('%Y%m%d'))
        ip_6day = []
        for t in time_:
            test = []
            while True:
                try:
                    wb = openpyxl.load_workbook("inputFile/" + t + "/IP_with_area.xlsx")
                    ws = wb.get_sheet_by_name("Sheet")
                    break
                except:
                    _xlxs_csv.Csv2Xlxs(t, "IP_with_area")
            for y in list(ws.rows):
                test.append(y[0].value)
            ip_6day.append(test)
            wb.close()
        count = 0
        for x in list(ws_source.rows):
            x = list(x)
            if count <1:
                title = []
                for y in x:
                    title.append(y.value)
                title.append(u'频率')
                ws_end.append(title)
                count += 1
                continue
            rate[x[0].value] = 1
            data = []
            for y in x:
                data.append(y.value)
            for day in ip_6day:
                for ip in day:
                    if ip == x[0].value:
                        rate[x[0].value] += 1
                        break
            data.append(rate[x[0].value])
            ws_end.append(data)
        wb_end.save("outputFile/" + time + u"/统计.xlsx")
        wb_end.close()
        wb_source.close()

    def area_Mate(self,time, class_):
        title = [u'IP', u'地区', u'次数', u'频率']
        wb_ = openpyxl.load_workbook("outputFile/" + time + "/Top5.xlsx")
        wb = openpyxl.load_workbook("outputFile/" + time + u"/统计.xlsx")
        ws_source = wb.get_sheet_by_name(u"IP(全)")
        ws_class = wb_.get_sheet_by_name(class_)
        wb_.create_sheet('temp')
        ws_temp = wb_.get_sheet_by_name('temp')
        ws_temp.append(title)
        count = 0
        for x in list(ws_class.rows):
            x = list(x)
            if count < 1:
                count += 1
                continue
            for y in list(ws_source.rows):
                y = list(y)
                if x[0].value == y[0].value:
                    data = []
                    data.append(y[0].value)
                    if y[2].value != u'中国':
                        data.append(y[2].value)
                    else:
                        str = ''
                        for m in (4, 5, 6):
                            if y[m].value != 'NULL':
                                str += y[m].value
                        data.append(str)
                    data.append(x[1].value)
                    data.append(y[7].value)
                    ws_temp.append(data)
                    break
        wb_.remove(ws_class)
        ws_temp.title = class_
        wb_.save("outputFile/" + time + "/Top5.xlsx")
        wb.close()