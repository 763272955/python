# -*- coding:utf-8 -*-

import _xlxs_csv
from _data_operat import Data_Operat as DO
import openpyxl
import re

class Daily_Operat(object):
    def __init__(self):
        self.top = {}
        self.charts = {}
        self.chartr = {}
        self.port = {}
        self.regular_type = []
        self.title = [u'告警时间', u'规则名称', u'源IP', u'源端口', u'目的IP',u'目的端口',
                      u'上报引擎', u'返回消息', u'网口编号', u'网口别名', u'全行']

    def internet_Event(self):
        self.__init__()
        filename = "Internet_Event"
        self.regular_type.append(re.compile(r'http.status_code=4(.*?);'))
        self.regular_type.append(re.compile(r'http.status_code=;'))
        self.regular_type.append(re.compile(r'http.url=/;'))
        self.regular_type.append(re.compile(r'http.url=;'))
        _xlxs_csv.Csv2Xlxs(self.time, filename)
        wb_source = openpyxl.load_workbook("inputFile/" + self.time + "/" + filename + ".xlsx")
        ws_source = wb_source.get_sheet_by_name("Sheet")
        ws_filter_portall = DO().get_Sheet(self.wb_filter, u"端口(全)", self.title)
        ws_filter_urlall = DO().get_Sheet(self.wb_filter, u"URL(全)", self.title)
        ws_chart_portall = DO().get_Sheet(self.wb_chart, u"端口(全)", [u"端口", u"次数"])
        ws_chart_urlall_sys = DO().get_Sheet(self.wb_chart, u"URL(全)系统", [u"系统", u"次数"])
        ws_chart_urlall_rul = DO().get_Sheet(self.wb_chart, u"URL(全)规则", [u"规则", u"次数"])
        ws_top_all = DO().get_Sheet(self.wb_top, u"URL(全)", [u"IP", u"次数"])
        for rows_source in list(ws_source.rows):
            rows_source = list(rows_source)
            sys_all = DO().system_Belong(rows_source[4].value, self.assetall)
            if sys_all != None:
                data = DO().get_Data(rows_source)
                data.append(sys_all)
                self.port = DO().count_Dict(self.port, data[5])
                ws_filter_portall.append(data)
                if data[7] != None:
                    if "http" in data[7]:
                        if DO().data_Regular(rows_source[7].value, self.regular_type):
                            ws_filter_urlall.append(data)
                            self.top = DO().count_Dict(self.top, data[2])
                            self.charts = DO().count_Dict(self.charts, data[-1])
                            self.chartr = DO().count_Dict(self.chartr, data[1])
        ws_top_all = DO().get_Dict_data(ws_top_all, self.top)
        ws_chart_urlall_sys = DO().get_Dict_data(ws_chart_urlall_sys, self.charts)
        ws_chart_urlall_rul = DO().get_Dict_data(ws_chart_urlall_rul, self.chartr)
        ws_chart_portall = DO().get_Dict_data(ws_chart_portall, self.port)
        self.wb_filter.save("outputFile/" + self.time + "/" + u"数据筛选" + ".xlsx")
        self.wb_chart.save("outputFile/" + self.time + "/" + u"图表" + ".xlsx")
        self.wb_top.save("outputFile/" + self.time + "/" + u"Top5" + ".xlsx")
        wb_source.close()
        print u"======原始筛选完成======"

    def vulnerability_Attack(self):
        filename = "Vulnerability_Attack"
        self.similar(filename, u"漏洞(全)")
        print u"======漏洞筛选完成======"

    def cross_Site(self):
        filename = "Cross_Site_Injection"
        self.similar(filename, u"跨站(全)")
        print u"======跨站筛选完成======"

    def login_Attempt(self):
        self.__init__()
        filename = "Login_Attempt"
        self.regular_type.append(re.compile(r'http.status_code=4(.*?);'))
        self.regular_type.append(re.compile(r'http.status_code=;'))
        _xlxs_csv.Csv2Xlxs(self.time, filename)
        wb_source = openpyxl.load_workbook("inputFile/" + self.time + "/" + filename + ".xlsx")
        ws_source = wb_source.get_sheet_by_name("Sheet")
        ws_filter_all = DO().get_Sheet(self.wb_filter, u"登录(全)", self.title)
        ws_chart_loginall = DO().get_Sheet(self.wb_chart, u"登录(全)", [u"系统", u"次数"])
        ws_top_all = DO().get_Sheet(self.wb_top, u"登录(全)", [u"IP", u"次数"])
        for rows_source in list(ws_source.rows):
            rows_source = list(rows_source)
            if DO().data_Regular(rows_source[7].value, self.regular_type):
                sys_all = DO().system_Belong(rows_source[4].value, self.assetall)
                if sys_all != None:
                    data = DO().get_Data(rows_source)
                    data.append(sys_all)
                    ws_filter_all.append(data)
                    self.top = DO().count_Dict(self.top, data[2])
                    self.charts = DO().count_Dict(self.charts, data[-1])
        ws_top_all = DO().get_Dict_data(ws_top_all, self.top)
        ws_chart_loginall = DO().get_Dict_data(ws_chart_loginall, self.charts)
        self.wb_filter.save("outputFile/" + self.time + "/" + u"数据筛选" + ".xlsx")
        self.wb_chart.save("outputFile/" + self.time + "/" + u"图表" + ".xlsx")
        self.wb_top.save("outputFile/" + self.time + "/" + u"Top5" + ".xlsx")
        wb_source.close()
        print u"======登录筛选完成======"

    def information_Detetion(self):
        filename = "Information_Detection"
        self.similar(filename, u"探测(全)")
        print u"======探测筛选完成======"

    def similar(self, filename, class_):
        self.__init__()
        self.regular_type.append(re.compile(r'http.status_code=4(.*?);'))
        self.regular_type.append(re.compile(r'http.status_code=;'))
        _xlxs_csv.Csv2Xlxs(self.time, filename)
        wb_source = openpyxl.load_workbook("inputFile/" + self.time + "/" + filename + ".xlsx")
        ws_source = wb_source.get_sheet_by_name("Sheet")
        ws_filter_all = DO().get_Sheet(self.wb_filter, class_, self.title)
        ws_chart_all_sys = DO().get_Sheet(self.wb_chart, class_+ u"系统", [u"系统", u"次数"])
        ws_chart_all_rul = DO().get_Sheet(self.wb_chart, class_ + u"规则", [u"规则", u"次数"])
        ws_top_all = DO().get_Sheet(self.wb_top, class_, [u"IP", u"次数"])
        for rows_source in list(ws_source.rows):
            rows_source = list(rows_source)
            if DO().data_Regular(rows_source[7].value, self.regular_type):
                sys_all = DO().system_Belong(rows_source[4].value, self.assetall)
                if sys_all != None:
                    data = DO().get_Data(rows_source)
                    data.append(sys_all)
                    ws_filter_all.append(data)
                    self.top = DO().count_Dict(self.top, data[2])
                    self.charts = DO().count_Dict(self.charts, data[-1])
                    self.chartr = DO().count_Dict(self.chartr, data[1])
        ws_top_all = DO().get_Dict_data(ws_top_all, self.top)
        ws_chart_all_sys = DO().get_Dict_data(ws_chart_all_sys, self.charts)
        ws_chart_all_rul = DO().get_Dict_data(ws_chart_all_rul, self.chartr)
        self.wb_filter.save("outputFile/" + self.time + "/" + u"数据筛选" + ".xlsx")
        self.wb_top.save("outputFile/" + self.time + "/" + u"Top5" + ".xlsx")
        self.wb_chart.save("outputFile/" + self.time + "/" + u"图表" + ".xlsx")
        wb_source.close()

    def IP(self):
        ws_source = self.wb_count.get_sheet_by_name(u"IP(全)")
        ws_char_country = DO().get_Sheet(self.wb_chart, u"IP(全)1", [u"国家"])
        ws_char_city = DO().get_Sheet(self.wb_chart, u"IP(全)2", [u"城市"])
        ws_top = DO().get_Sheet(self.wb_top, u"IP(全)", [u"IP", u"次数"])
        count = 0
        data = []
        data_sort = []
        for row in list(ws_source.rows):
            data_ = []
            row = list(row)
            if count < 1:
                count += 1
                continue
            for x in row:
                data_.append(x.value)
            data.append(data_)
        for x in range(len(data)):
            if x == 0:
                data_sort.append(data[x])
                continue
            data_sort.append(data[x])
            for i in range(1, x+1):
                if int(data_sort[x+1-i][1]) > int(data_sort[x - i][1]):
                    a = data_sort[x-i]
                    data_sort[x-i] = data_sort[x+1-i]
                    data_sort[x+1-i] = a
        for data_ in data_sort:
            ws_char_country.append([data_[2]])
            if u"澳门" not in data_[4] and u"香港" not in data_[4] and u"台湾" not in data_[4] and u"NULL" not in data_[4] and u"中国" in data_[2]:
                ws_char_city.append([data_[4]])
            ws_top.append([data_[0], data_[1]])
        self.wb_top.save("outputFile/" + self.time + "/" + u"Top5" + ".xlsx")
        self.wb_chart.save("outputFile/" + self.time + "/" + u"图表" + ".xlsx")

    def rate_Count(self):
        filename = 'IP_with_area'
        _xlxs_csv.Csv2Xlxs(self.time, filename)
        DO().rate_Count(self.time)
        print u"======频率统计完成======"

    def open_Excel(self):
        filtername = [u"端口(全)", u"URL(全)", u"漏洞(全)", u"跨站(全)", u"登录(全)", u"探测(全)"]
        chartname = [u"IP(全)1", u"IP(全)2", u"端口(全)", u"URL(全)系统", u"URL(全)规则", u"漏洞(全)系统", u"漏洞(全)规则",
                     u"跨站(全)系统", u"跨站(全)规则", u"登录(全)", u"探测(全)系统",u"探测(全)规则"]
        topname = [u"IP(全)", u"URL(全)", u"漏洞(全)", u"跨站(全)", u"登录(全)", u"探测(全)"]
        countname = [u"统计"]
        self.wb_filter = DO().create_Newsheet("outputFile/" + self.time + "/" + u"数据筛选" + ".xlsx", filtername)
        self.wb_chart = DO().create_Newsheet("outputFile/" + self.time + "/" + u"图表" + ".xlsx", chartname)
        self.wb_top = DO().create_Newsheet("outputFile/" + self.time + "/" + u"Top5" + ".xlsx", topname)
        self.wb_count = DO().create_Newsheet("outputFile/" + self.time + "/" + u"统计" + ".xlsx", countname)

    def get_Asset(self):
        wb_asset = openpyxl.load_workbook('inputFile/assets2017-5-26.xlsx')
        ws_assetall = wb_asset.get_sheet_by_name(u"全行资产")
        self.assetall = []
        count = 0
        for x in list(ws_assetall.rows):
            if count < 1:
                count += 1
                continue
            self.assetall.append([x[1].value, x[2].value])

    def data_Count_(self, ws_count, x):
        number = len(list(self.wb_filter.get_sheet_by_name(x + u'(全)').rows)) - 1
        number_system = len(list(self.wb_chart.get_sheet_by_name(x + u'(全)系统').rows)) - 1
        if number_system != 0:
            system = list(list(self.wb_chart.get_sheet_by_name(x + u'(全)系统').rows)[1])[0].value
        else:
            system = None
        if number != 0:
            percent_system = float(
                list(list(self.wb_chart.get_sheet_by_name(x + u'(全)系统').rows)[1])[1].value / float(number))
        else:
            percent_system = None
        if number_system != 0:
            rule = list(list(self.wb_chart.get_sheet_by_name(x + u'(全)规则').rows)[1])[0].value
        else:
            rule = None
        ws_count.append([x, number, number_system, system, percent_system, rule])
        return ws_count

    def data_Count(self):
        ws_count = DO().get_Sheet(self.wb_count, u"统计", [u"种类", u"次数", u"系统个数", u"系统", u"百分比", u"规则"])
        ws_count.append([u"告警总数", len(list(self.wb_filter.get_sheet_by_name(u'端口(全)').rows))-1])
        ws_count.append([u"IP", len(list(self.wb_top.get_sheet_by_name(u'IP(全)').rows)) - 1])
        ws_count.append([u"端口", len(list(self.wb_chart.get_sheet_by_name(u'端口(全)').rows)) - 1])
        ws_count = self.data_Count_(ws_count, u"URL")
        ws_count = self.data_Count_(ws_count, u"漏洞")
        ws_count = self.data_Count_(ws_count, u"跨站")
        ws_count.append([u"登录(全)", len(list(self.wb_filter.get_sheet_by_name(u'登录(全)').rows))-1])
        ws_count = self.data_Count_(ws_count, u"探测")
        self.wb_count.save("outputFile/" + self.time + "/" + u"统计" + ".xlsx")
        print u"======数据统计完成======"

    def top5_Count(self):
        class_ = [u"IP(全)", u"URL(全)", u"漏洞(全)", u"跨站(全)", u"登录(全)", u"探测(全)", ]
        for x in class_:
            DO().area_Mate(self.time, x)
        print u"======Top5统计完成======"

    def run(self, time):
        self.time = time
        self.rate_Count()
        self.get_Asset()
        self.open_Excel()
        self.internet_Event()
        self.vulnerability_Attack()
        self.cross_Site()
        self.login_Attempt()
        self.information_Detetion()
        self.IP()
        self.top5_Count()
        self.data_Count()