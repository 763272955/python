# -*- coding:utf-8 -*-

import openpyxl
import datetime
import re

class Data_Operat(object):
    def __init__(self):
        self.sys = None

    def delete_List(self, data, index):
        data = data[:index] + data[index+1:]
        return data

    def create_Newsheet(self, filename, sheetname):
        wb = openpyxl.load_workbook(filename)
        for name in sheetname:
            wb.create_sheet(name)
        try:
            wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        except:
            pass
        return wb

    def get_Sheet(self, wb, name, title):
        ws = wb.get_sheet_by_name(name)
        ws.append(title)
        return ws

    def save_Excel(self, wb, filename):
        wb.save(filename)
        wb.close()

    def get_Data(self, rows):
        data = []
        for row in rows:
            data.append(row.value)
        return data

    def count_Dict(self, dict, key, data_=[]):
        if key not in dict.keys():
            dict[key] = []
            dict[key].append(1)
            for x in data_:
                dict[key].append(x)
        else:
            dict[key][0] += 1
            for i in range(len(data_)):
                if data_[i] not in dict[key][i+1]:
                    dict[key][i+1] += ',' + data_[i]
        return dict

    def get_Dict_data(self, ws=None, dict=None, des=None, delet=None, rechange=None):
        data = []
        data_sort = []
        for key in dict.keys():
            data_ = []
            data_.append(key)
            for x in dict[key]:
                data_.append(x)
            data.append(data_)
        for x in range(len(data)):
            if x == 0:
                data_sort.append(data[x])
                continue
            data_sort.append(data[x])
            for i in range(1, x+1):
                if int(data_sort[x+1-i][des]) > int(data_sort[x - i][des]):
                    a = data_sort[x-i]
                    data_sort[x-i] = data_sort[x+1-i]
                    data_sort[x+1-i] = a
        if delet != None:
            data_sort_ = []
            for x in data_sort:
                data_sort_.append(self.delete_List(x, delet))
            data_sort = data_sort_
        if rechange != None:
            data_sort_ = []
            for x in data_sort:
                m = x[rechange[0]]
                data_x = self.delete_List(x, rechange[0])
                data_x.insert(rechange[1], m)
                data_sort_.append(data_x)
            data_sort = data_sort_
        for x in range(len(data_sort)):
            ws.append(data_sort[x])
        return ws

    def http_Split(self, response):
        host = re.compile(r'http.host=(.*?);').findall(response)
        url = re.compile(r'http.url=(.*?);').findall(response)
        user_agent = re.compile(r'http.user_agent=(.*?);').findall(response)
        status_code = re.compile(r'http.status_code=(.*?);').findall(response)
        if len(host) == 0:
            host = "NULL"
        else:
            host = host[0]
        if len(url) == 0:
            url = "NULL"
        else:
            url = url[0]
        if len(user_agent) == 0:
            user_agent = "NULL"
        else:
            user_agent = user_agent[0]
        if len(status_code) == 0:
            status_code = "NULL"
        else:
            status_code = status_code[0]
        return host, url, user_agent, status_code

    def system_Belong(self, ip, asset):
        for a in asset:
            if ip == a[0]:
                self.sys = a[1]
        return self.sys

    def branch_Belong(self, ip, asset_lan):
        ip_split = ip.split('.')
        branch = u"未知分行"
        equipment = u"未知设备"
        for data in asset_lan:
            data_S_split = data[-2].split('.')
            data_O_split = data[-1].split('.')
            if ip_split[0:3] == data_S_split[0:3]:
                if int(ip_split[-1]) >= int(data_S_split[-1]) and int(ip_split[-1]) <= int(data_O_split[-1]):
                    branch = data[0]
                    equipment = data[1]
        return branch, equipment

    def data_Regular(self, response, record_type, flag=0):
        if response != None:
            for type in record_type:
                result = type.findall(response)
                if flag == 1:
                    if len(result) != 0:
                        return False
                if flag == 0:
                    if len(result) != 0:
                        if result[0].strip() == '':
                            return False
        return True

    def rate_Count(self, time):
        time_ = []
        rate = {}
        wb_source = openpyxl.load_workbook("inputFile/" + time + "/IP_with_area.xlsx")
        wb_end = openpyxl.load_workbook("outputFile/" + time + u"/统计.xlsx")
        wb_end.create_sheet(u"IP(全)")
        ws_source = wb_source.get_sheet_by_name("Sheet")
        ws_end = wb_end.get_sheet_by_name(u"IP(全)")
        for i in range(2, 8):
            time_.append((datetime.datetime.now() - datetime.timedelta(days=i)).strftime('%Y%m%d'))
        ip_6day = []
        for t in time_:
            test = []
            wb = openpyxl.load_workbook("inputFile/" + t + "/IP_with_area.xlsx")
            ws = wb.get_sheet_by_name("Sheet")
            for y in list(ws.rows):
                test.append(y[0].value)
            ip_6day.append(test)
            wb.close()
        count = 0
        for x in list(ws_source.rows):
            x = list(x)
            if count <1:
                title = []
                for y in x:
                    title.append(y.value)
                title.append(u'频率')
                ws_end.append(title)
                count += 1
                continue
            rate[x[0].value] = 1
            data = []
            for y in x:
                data.append(y.value)
            for day in ip_6day:
                for ip in day:
                    if ip == x[0].value:
                        rate[x[0].value] += 1
                        break
            data.append(rate[x[0].value])
            ws_end.append(data)
        wb_end.save("outputFile/" + time + u"/统计.xlsx")
        wb_end.close()
        wb_source.close()

    def area_Mate(self,time, class_):
        title = [u'IP', u'地区', u'次数', u'频率']
        wb_ = openpyxl.load_workbook("outputFile/" + time + "/Top5.xlsx")
        wb = openpyxl.load_workbook("outputFile/" + time + u"/统计.xlsx")
        ws_source = wb.get_sheet_by_name(u"IP(全)")
        ws_class = wb_.get_sheet_by_name(class_)
        wb_.create_sheet('temp')
        ws_temp = wb_.get_sheet_by_name('temp')
        ws_temp.append(title)
        count = 0
        for x in list(ws_class.rows):
            x = list(x)
            if count < 1:
                count += 1
                continue
            for y in list(ws_source.rows):
                y = list(y)
                if x[0].value == y[0].value:
                    data = []
                    data.append(y[0].value)
                    if y[2].value != u'中国':
                        data.append(y[2].value)
                    else:
                        str = ''
                        for m in (4, 5, 6):
                            if y[m].value != 'NULL':
                                str += y[m].value
                        data.append(str)
                    data.append(x[1].value)
                    data.append(y[7].value)
                    ws_temp.append(data)
                    break
        wb_.remove(ws_class)
        ws_temp.title = class_
        wb_.save("outputFile/" + time + "/Top5.xlsx")
        wb.close()