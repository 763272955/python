# -*- coding:utf-8 -*-

import _xlxs_csv
from _data_operat import Data_Operat as DO
import openpyxl
import re

class Daily_Operat(object):
    def __init__(self):
        self.topall = {}
        self.topzz = {}
        self.topwy = {}
        self.chartall = {}
        self.chartzz = {}
        self.charts = {}
        self.chartwy = {}
        self.regular_type_1 = []
        self.regular_type_0 = []
        self.title = [u'告警时间', u'规则名称', u'源IP', u'源端口', u'目的IP', u'目的端口',
                      u'上报引擎', u'返回消息', u'网口编号', u'网口别名', u'分行']
        self.title_f = [u'告警时间', u'规则名称', u'源IP', u'源端口', u'目的IP',u'目的端口',
                      u'上报引擎', u'Host', u"Url", u"User-Agent", u"Status_Code",
                      u'网口编号', u'网口别名', u'分行']
        self.title_lan = [u'告警时间', u'规则名称', u'源IP', u'源端口', u'目的IP', u'目的端口',
                      u'上报引擎', u'返回消息',u'网口编号', u'网口别名', u'源所属分行', u"源设备",
                          u"目所属分行", u"目设备"]

    def internet_Event(self):
        self.__init__()
        filename = "Internet_Event"
        self.regular_type_1.append(re.compile(r'http.status_code=4(.*?);'))
        self.regular_type_0.append(re.compile(r'http.status_code=(.*?);'))
        self.regular_type_0.append(re.compile(r'http.url=/(.*?);'))
        self.regular_type_0.append(re.compile(r'http.url=(.*?);'))
        try:
            _xlxs_csv.Csv2Xlxs(self.time, filename)
        except:
            pass
        wb_source = openpyxl.load_workbook("inputFile/" + self.time + "/" + filename + ".xlsx")
        ws_source = wb_source.get_sheet_by_name("Sheet")
        ws_filter_urlall = DO().get_Sheet(self.wb_filter, u"URL(分)", self.title)
        ws_filter_url_f = DO().get_Sheet(self.wb_filter, u"URL(分)筛", self.title_f)
        ws_chart_urlsys = DO().get_Sheet(self.wb_chart, u"URL(分)系统", [u"系统", u"次数"])
        ws_chart_url = DO().get_Sheet(self.wb_chart, u"URL(分)筛URL", [u"URL", u"次数"])
        ws_top_url = DO().get_Sheet(self.wb_top, u"URL(分)", [u"URL", u"应用系统", u"源IP", u"告警名称", u"频率"])
        for rows_source in list(ws_source.rows):
            rows_source = list(rows_source)
            sys_other = DO().system_Belong(rows_source[4].value, self.assetother)
            if sys_other != None:
                data = DO().get_Data(rows_source)
                data.append(sys_other)
                ws_filter_urlall.append(data)
                self.charts = DO().count_Dict(self.charts, data[-1])
                if rows_source[7].value != None:
                    if "http.host" in rows_source[7].value:
                        if DO().data_Regular(rows_source[7].value, self.regular_type_1, 1):
                            if DO().data_Regular(rows_source[7].value, self.regular_type_0, 0):
                                host, url, user_agent, status_code = DO().http_Split(data[7])
                                for x in status_code, user_agent, url, host:
                                    data.insert(7, x)
                                ws_filter_url_f.append(data)
                                if host != 'NULL':
                                    self.chartall = DO().count_Dict(self.chartall, host + data[8])
                                    self.topall = DO().count_Dict(self.topall, host + data[8], [data[-1], data[2], data[1]])
                                else:
                                    self.chartall = DO().count_Dict(self.chartall, data[4] + data[8])
                                    self.topall = DO().count_Dict(self.topall, data[4] + data[8], [data[-1], data[2], data[1]])
        ws_chart_url = DO().get_Dict_data(ws=ws_chart_url, dict=self.chartall, des=1)
        ws_chart_urlsys = DO().get_Dict_data(ws=ws_chart_urlsys, dict=self.charts, des=1)
        ws_top_url = DO().get_Dict_data(ws=ws_top_url, dict=self.topall, des=1, delet=1)
        self.wb_filter.save("outputFile/" + self.time + "/" + u"数据筛选" + ".xlsx")
        self.wb_chart.save("outputFile/" + self.time + "/" + u"图表" + ".xlsx")
        self.wb_top.save("outputFile/" + self.time + "/" + u"Top" + ".xlsx")
        wb_source.close()
        print u"======分行筛选完成======"

    def lan_Event(self):
        self.__init__()
        filename = "Lan_Event"
        self.regular_type_1.append(re.compile(u"r'http.status_code=4(.*?);'"))
        self.regular_type_0.append(re.compile(u"r'http.status_code=(.*?);'"))
        try:
            _xlxs_csv.Csv2Xlxs(self.time, filename)
        except:
            pass
        wb_source = openpyxl.load_workbook("inputFile/" + self.time + "/" + filename + ".xlsx")
        ws_source = wb_source.get_sheet_by_name("Sheet")
        ws_filter_lan = DO().get_Sheet(self.wb_filter, u"内网(分)", self.title_lan)
        ws_filter_zz = DO().get_Sheet(self.wb_filter, u"自助设备", self.title_lan)
        ws_filter_wy = DO().get_Sheet(self.wb_filter, u"网银体验", self.title_lan)
        ws_chart_lan = DO().get_Sheet(self.wb_chart, u"内网(分)", [u"分行", u"次数"])
        ws_chart_zz = DO().get_Sheet(self.wb_chart, u"自助设备", [u"分行", u"次数"])
        ws_chart_wy = DO().get_Sheet(self.wb_chart, u"网银体验", [u"分行", u"次数"])
        ws_top_lan = DO().get_Sheet(self.wb_top, u"内网(分)", [u"源IP", u"源所属分行/设备", u"目IP", u"目所属分行/设备", u"规则", u"次数"])
        ws_top_zz = DO().get_Sheet(self.wb_top, u"自助设备", [u"源IP", u"源所属分行/设备", u"目IP", u"目所属分行/设备", u"规则", u"次数"])
        ws_top_wy = DO().get_Sheet(self.wb_top, u"网银体验", [u"源IP", u"源所属分行/设备", u"目IP", u"目所属分行/设备", u"规则", u"次数"])
        for rows_source in list(ws_source.rows):
            rows_source = list(rows_source)
            if u"告警时间" not in rows_source[0].value:
                if DO().data_Regular(rows_source[7].value, self.regular_type_1, 1):
                    if DO().data_Regular(rows_source[7].value, self.regular_type_0, 0):
                        sbranch, sequ = DO().branch_Belong(rows_source[2].value, self.assetlan)
                        obranch, oequ = DO().branch_Belong(rows_source[4].value, self.assetlan)
                        data = DO().get_Data(rows_source)
                        data.append(sbranch)
                        data.append(sequ)
                        data.append(obranch)
                        data.append(oequ)
                        ws_filter_lan.append(data)
                        self.chartall = DO().count_Dict(self.chartall, data[-2])
                        self.topall = DO().count_Dict(self.topall, data[2], [data[-4]+'/'+data[-3], data[4], data[-2]+'/'+data[-1], data[1]])
                        if u"自助设备" in sequ or u"自助设备" in oequ:
                            ws_filter_zz.append(data)
                            self.chartzz = DO().count_Dict(self.chartzz, data[-2])
                            self.topzz = DO().count_Dict(self.topzz, data[2], [data[-4]+'/'+data[-3], data[4], data[-2]+'/'+data[-1], data[1]])
                        if u"网银体验" in sequ or u"网银体验" in oequ:
                            ws_filter_wy.append(data)
                            self.chartwy = DO().count_Dict(self.chartwy, data[-2])
                            self.topwy = DO().count_Dict(self.topwy, data[2], [data[-4]+'/'+data[-3], data[4], data[-2]+'/'+data[-1], data[1]])
        ws_chart_lan = DO().get_Dict_data(ws_chart_lan, self.chartall, 1)
        ws_chart_zz = DO().get_Dict_data(ws_chart_zz, self.chartzz, 1)
        ws_chart_wy = DO().get_Dict_data(ws_chart_wy, self.chartwy, 1)
        ws_top_lan = DO().get_Dict_data(ws_top_lan, self.topall, 1)
        ws_top_zz = DO().get_Dict_data(ws_top_zz, self.topzz, 1)
        ws_top_wy = DO().get_Dict_data(ws_top_wy, self.topwy, 1)
        self.wb_filter.save("outputFile/" + self.time + "/" + u"数据筛选" + ".xlsx")
        self.wb_chart.save("outputFile/" + self.time + "/" + u"图表" + ".xlsx")
        self.wb_top.save("outputFile/" + self.time + "/" + u"Top" + ".xlsx")
        wb_source.close()
        print u"======内网筛选完成======"

    def open_Excel(self):
        filtername = [u"URL(分)", u"URL(分)筛", u"内网(分)", u"自助设备", u"网银体验"]
        chartname = [u"URL(分)系统",  u"URL(分)筛URL",  u"内网(分)", u"自助设备", u"网银体验"]
        topname = [u"URL(分)", u"内网(分)", u"自助设备", u"网银体验"]
        countname = [u"统计"]
        self.wb_filter = DO().create_Newsheet("outputFile/" + self.time + "/" + u"数据筛选" + ".xlsx", filtername)
        self.wb_chart = DO().create_Newsheet("outputFile/" + self.time + "/" +u"图表" + ".xlsx", chartname)
        self.wb_top = DO().create_Newsheet("outputFile/" + self.time + "/" + u"Top" + ".xlsx", topname)
        self.wb_count = DO().create_Newsheet("outputFile/" + self.time + "/" + u"统计" + ".xlsx", countname)
        self.wb_rate_internet = DO().create_Newsheet("../IP/Internet.xlsx", [self.time])
        self.wb_rate_lan = DO().create_Newsheet("../IP/Lan.xlsx", [self.time])

    def data_Count(self):
        ws_count = DO().get_Sheet(self.wb_count, u"统计", [u"类别", u"次数"])
        ws_count.append([u"互联网告警数", len(list(self.wb_filter.get_sheet_by_name(u"URL(分)").rows))-1])
        ws_count.append([u"内网告警数", len(list(self.wb_filter.get_sheet_by_name(u"内网(分)")))-1])
        ws_count.append([u"自助设备数", len(list(self.wb_filter.get_sheet_by_name(u"自助设备")))-1])
        ws_count.append([u"网银体验数", len(list(self.wb_filter.get_sheet_by_name(u"网银体验")))-1])
        self.wb_count.save("outputFile/" + self.time + "/" + u"统计" + ".xlsx")

    def get_Asset(self):
        wb_asset = openpyxl.load_workbook('inputFile/assets2017-5-26.xlsx')
        wb_asset_lan = openpyxl.load_workbook('inputFile/all.xlsx')
        wb_asset_2to1 = openpyxl.load_workbook('inputFile/2to1.xlsx')
        ws_assetother = wb_asset.get_sheet_by_name(u"分行资产")
        ws_asset_lan = wb_asset_lan.get_sheet_by_name(u"Sheet")
        ws_asset_2to1 = wb_asset_2to1.get_sheet_by_name(u"Sheet")
        self.assetother = []
        self.assetlan = []
        self.asset2to1 = []
        count = 0
        for x in list(ws_assetother.rows):
            if count < 1:
                count += 1
                continue
            self.assetother.append([x[1].value, x[2].value])
        for x in list(ws_asset_2to1.rows):
            data = DO().get_Data(x)
            self.asset2to1.append(data)
        for x in list(ws_asset_lan.rows):
            if x[6].value != None:
                result = re.compile(r'\((.*?)\)').findall(x[1].value)[0]
                for y in self.asset2to1:
                    if result in y:
                        branch = y[0]
                self.assetlan.append([branch, x[4].value, x[6].value, x[7].value])

    def run(self, time):
        self.time = time
        self.get_Asset()
        self.open_Excel()
        self.internet_Event()
        self.lan_Event()
        self.data_Count()
        self.wb_count.close()
        self.wb_filter.close()
        self.wb_top.close()
        self.wb_chart.close()