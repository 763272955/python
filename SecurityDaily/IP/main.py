# -*- coding:utf-8 -*-

import openpyxl
import datetime
import csv

def Csv2Xlxs(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    f = open(filename + ".csv")
    reader = csv.reader(f, delimiter=',')
    for row in reader:
        row_container = []
        for cell in row:
            try:
                cell = int(float(cell))
                row_container.append(cell)
                continue
            except:
                pass
            row_container.append(cell.decode('gbk').encode('utf-8'))
        ws.append(row_container)
    f.close()
    wb.save(filename + ".xlsx")

def create_Newsheet(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    for name in sheetname:
        try:
            wb.remove_sheet(wb.get_sheet_by_name(name))
        except:
            pass
        wb.create_sheet(name)
    try:
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    except:
        pass
    return wb

def get_Sheet(wb, name, title):
    ws = wb.get_sheet_by_name(name)
    ws.append(title)
    return ws

if __name__ == '__main__':
    Csv2Xlxs("Internet_Event")
    Csv2Xlxs("Lan_Event")
    time_ = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime('%Y%m%d')
    wb_source_internet = openpyxl.load_workbook("Internet_Event.xlsx")
    wb_source_lan = openpyxl.load_workbook("Lan_Event.xlsx")
    wb_internet = create_Newsheet("Internet.xlsx", [time_])
    wb_lan = create_Newsheet("Lan.xlsx", [time_])
    ws_source_internet = wb_source_internet.get_sheet_by_name("Sheet")
    ws_source_lan = wb_source_lan.get_sheet_by_name("Sheet")
    ws_internet = get_Sheet(wb_internet, time_, ["IP", u"频率"])
    ws_lan = get_Sheet(wb_lan, time_, ["IP", u"频率"])
    internet_day = []
    lan_day = []
    for x in range(2, 8):
        time__ = (datetime.datetime.now()-datetime.timedelta(days=x)).strftime('%Y%m%d')
        try:
            ws_i = wb_internet.get_sheet_by_name(time__)
        except:
            print u"未找到 %s 的IP,请手动添加!" % time__
            print u"添加路径,D:/SecurityDaily/IP/Internet.xlsx"
            raise exit()
        try:
            ws_l = wb_lan.get_sheet_by_name(time__)
        except:
            print u"未找到 %s 的IP,请手动添加!" % time__
            print u"添加路径,D:/SecurityDaily/IP/Lan.xlsx"
            raise exit()
        test = []
        for row in list(ws_i.rows):
            if 'IP' not in row[0].value:
                test.append(row[0].value)
        internet_day.append(test)
        test = []
        for row in list(ws_l.rows):
            if 'IP' not in row[0].value:
                test.append(row[0].value)
        lan_day.append(test)
    test = []
    for internet in list(ws_source_internet.rows):
        if internet[2].value in test:
            continue
        test.append(internet[2].value)
        count = 1
        if "IP" not in internet[2].value:
            for x in internet_day:
                if internet[2].value in x:
                    count += 1
            ws_internet.append([internet[2].value, count])
    test = []
    for lan in list(ws_source_lan.rows):
        count = 1
        if internet[2].value in test:
            continue
        test.append(internet[2].value)
        if "IP" not in lan[2].value:
            print lan[2].value
            for x in lan_day:
                if lan[2].value in x:
                    count += 1
            ws_lan.append([lan[2].value, count])
    wb_internet.save("Internet.xlsx")
    wb_lan.save("Lan.xlsx")
    wb_lan.close()
    wb_internet.close()
    wb_source_internet.close()
    wb_source_lan.close()