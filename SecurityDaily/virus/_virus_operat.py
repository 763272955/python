# -*- coding:utf-8 -*-

import openpyxl
import re
from _data_operat import Data_Operat as DO

class Virus_Operat(object):
    def __init__(self):
        self.time = None
        self.wb_virus = None
        self.wb_virus_source = None
        self.asset_branch = []
        self.asset_virus = []

    def asset_Get(self):
        asset_2to1 = []
        wb_asset_virus = openpyxl.load_workbook("assetFile/virus_asset.xlsx")
        wb_asset_branch = openpyxl.load_workbook("assetFile/branch_asset.xlsx")
        wb_asset_2to1 = openpyxl.load_workbook("assetFile/2to1.xlsx")
        ws_asset_virus = wb_asset_virus.get_sheet_by_name("Sheet")
        ws_asset_branch = wb_asset_branch.get_sheet_by_name("Sheet")
        ws_asset_2to1 = wb_asset_2to1.get_sheet_by_name("Sheet")
        for virus in list(ws_asset_virus.rows):
            self.asset_virus.append([virus[0].value, virus[1].value])
        for row in list(ws_asset_2to1.rows):
            data = DO().get_Data(row)
            asset_2to1.append(data)
        for branch in list(ws_asset_branch.rows):
            if branch[6].value != None:
                result = re.compile(r'\((.*?)\)').findall(branch[1].value)[0]
                for branch_ in asset_2to1:
                    if result in branch_:
                        result = branch_[0]
                    self.asset_branch.append([result, branch[4].value, branch[6].value, branch[7].value])
        print u"======资产获取完毕======"

    def excel_Open(self):
        self.wb_virus_source = openpyxl.load_workbook("inputFile/" + self.time+ ".xlsx")
        self.wb_virus = DO().create_Newsheet(excel_name="outputFile/" + self.time+ ".xlsx", sheet_name=[u"日志", u"主机类型", u"病毒类型",
                                                                                                                u"攻击类型", u"感染类型"])

    def virus_Operat(self):
        dict_host = {}
        dict_U = {}
        dict_virus = {}
        dict_operat = {}
        ws_virus_source = self.wb_virus_source.get_sheet_by_name("Sheet")
        ws_virus_daily = DO().get_Sheet(wb=self.wb_virus, sheet_name=u"日志", title=[u"结构", u"IP地址", u"主机类型", u"MAC地址",
                                                                                   u"计算机名",	u"病毒名称", u"病毒类型", u"受感染文件",
                                                                                   u"感染路径", u"攻击类型", u"处理措施", u"感染类型",
                                                                                   u"时间", u"扫描类型",u"组件版本", u"操作系统"])
        ws_virus_host = DO().get_Sheet(wb=self.wb_virus, sheet_name=u"主机类型", title=[u"主机类型", u"受攻击次数", u"所占比例"])
        ws_virus_virus = DO().get_Sheet(wb=self.wb_virus, sheet_name=u"病毒类型", title=[u"病毒类型", u"所占次数", u"所占比例"])
        ws_virus_U = DO().get_Sheet(wb=self.wb_virus, sheet_name=u"攻击类型", title=[u"攻击类型", u"所占次数", u"所占比例"])
        ws_virus_operat = DO().get_Sheet(wb=self.wb_virus, sheet_name=u"感染类型", title=[u"感染类型", u"所占次数", u"所占比例"])
        for source in list(ws_virus_source.rows):
            if u"IP地址" == source[1].value:
                continue
            data = DO().get_Data(source)
            host = DO().system_Belong(ip=source[1].value, asset=self.asset_branch, defaulthost=u"未知设备")
            data.insert(2, host)
            dict_host = DO().dict_Count(dict=dict_host, key=host)
            virus = DO().virus_Belong(virus=source[4].value, asset=self.asset_virus, defaultvirus=u"未知病毒")
            data.insert(6, virus)
            dict_virus = DO().dict_Count(dict=dict_virus, key=virus)
            U = DO().U_Belong(U=source[6].value)
            data.insert(9, U)
            dict_U = DO().dict_Count(dict=dict_U, key=U)
            operat = DO().operat_Belong(operat=source[7].value)
            data.insert(11, operat)
            dict_operat = DO().dict_Count(dict=dict_operat, key=operat)
            ws_virus_daily.append(data)
        print u"======日志统计完毕======"
        ws_virus_host = DO().dict_Getdata(ws=ws_virus_host, dict=dict_host)
        print u"======主机统计完毕======"
        ws_virus_virus = DO().dict_Getdata(ws=ws_virus_virus, dict=dict_virus)
        print u"======病毒统计完毕======"
        ws_virus_U = DO().dict_Getdata(ws=ws_virus_U, dict=dict_U)
        print u"======攻击统计完毕======"
        ws_virus_operat = DO().dict_Getdata(ws=ws_virus_operat, dict=dict_operat)
        print u"======感染统计完毕======"
        self.wb_virus.save("outputFile/" + self.time+ ".xlsx")

    def run(self, time_):
        self.time = time_
        self.asset_Get()
        self.excel_Open()
        self.virus_Operat()
        self.wb_virus.close()
        self.wb_virus_source.close()