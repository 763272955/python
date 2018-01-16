# -*- coding:utf-8 -*-

from _data_operat import Data_Operat as DO
from _virus_operat import Virus_Operat as VO
from _excel_create import Excel_Creat as EC
from _sql_conn import Sql_Conn as SC
from _alarm_detect import Alarm_Detect
from _color import Color
from _log_virus import Log_Virus as LV
import datetime
import time
import os
import openpyxl
import re

def asset_Get():
    asset_virus = []
    asset_branch = []
    asset_2to1 = []
    wb_asset_virus = openpyxl.load_workbook("assetFile/virus_asset.xlsx")
    wb_asset_branch = openpyxl.load_workbook("assetFile/branch_asset.xlsx")
    wb_asset_2to1 = openpyxl.load_workbook("assetFile/2to1.xlsx")
    ws_asset_virus = wb_asset_virus.get_sheet_by_name("Sheet")
    ws_asset_branch = wb_asset_branch.get_sheet_by_name("Sheet")
    ws_asset_2to1 = wb_asset_2to1.get_sheet_by_name("Sheet")
    for virus in list(ws_asset_virus.rows):
        asset_virus.append([virus[0].value, virus[1].value])
    for row in list(ws_asset_2to1.rows):
        data = DO().get_Data(row)
        asset_2to1.append(data)
    for branch in list(ws_asset_branch.rows):
        if branch[6].value != None:
            result = re.compile(r'\((.*?)\)').findall(branch[1].value)[0]
            for branch_ in asset_2to1:
                if result in branch_:
                    result = branch_[0]
                asset_branch.append([result, branch[3].value, branch[4].value, branch[6].value, branch[7].value, branch[8].value])
    print u"======资产获取完毕======"
    return asset_virus, asset_branch

if __name__ == '__main__':
    time_start = time.time()
    host = ''
    user = ''
    pwd = ''
    db = ''
    getTime_sql = "select top 1 clf_logReceivedtime " \
                  "from tb_avviruslog inner join tb_EntityInfo on tb_avviruslog.vlf_ClientGUID = tb_EntityInfo.EI_entityID " \
                  "where tb_avviruslog.VLF_FileName not like '%损害清除%' " \
                  "and tb_avviruslog.VLF_FileName not like '%DCS%' " \
                  "order by clf_logReceivedtime desc"
    end_ = SC(host, user, pwd, db).sql_Act(getTime_sql)
    asset_virus, asset_branch = asset_Get()
    while True:
        start_ = end_
        time.sleep(300)
        conn, cur = SC(host, user, pwd, db).get_Conn()
        cur.execute(getTime_sql)
        end_ = cur.fetchall()
        print u"::::::::::::::::::::::::::::::::::::::::"
        print u"::::::::::: %s :::::::::::" % datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        print u"::::::::::::::::::::::::::::::::::::::::"
        if start_[0][0] != end_[0][0]:
            sql = "select " \
                  "tb_avviruslog.clf_logReceivedtime, " \
                  "tb_avviruslog.clf_computername," \
                  "tb_EntityInfo.EI_ipAddresslist," \
                  "tb_EntityInfo.EI_MACAddressList," \
                  "tb_avviruslog.vlf_infectiondestination," \
                  "tb_avviruslog.vlf_virusname," \
                  "tb_avviruslog.vlf_filename," \
                  "tb_avviruslog.VLF_InfectionSource, " \
                  "tb_avviruslog.vlf_filepath," \
                  "tb_avviruslog.vlf_firstactionresult," \
                  "tb_avviruslog.clf_loggenerationtime," \
                  "tb_avviruslog.vlf_functioncode," \
                  "tb_avviruslog.vlf_patternnumber, " \
                  "tb_EntityInfo.EI_OS_Name " \
                  "from tb_avviruslog inner join tb_EntityInfo on tb_avviruslog.vlf_ClientGUID = tb_EntityInfo.EI_entityID " \
                  "where tb_avviruslog.clf_loggenerationtime > '" + str(start_[0][0]) + "' " \
                  "and tb_avviruslog.clf_loggenerationtime <= '" + str(end_[0][0]) + "' " \
                  "and tb_avviruslog.VLF_FileName not like '%损害清除%' " \
                  "and tb_avviruslog.VLF_FileName not like '%DCS%'"
            cur.execute(sql)
            retlist = cur.fetchall()
            flag = Alarm_Detect().detect(retlist, asset_branch, asset_virus)
            if flag == False:
                Color().print_green_text(u"未检测到告警")
        else:
            Color().print_green_text(u"未录入新数据")