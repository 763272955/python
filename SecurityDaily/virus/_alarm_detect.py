# -*- coding:utf-8 -*-

from _color import Color
from _data_operat import Data_Operat
import datetime
import os
import openpyxl

class Alarm_Detect(object):
    def __init__(self):
        self.clr = Color()
        self.date = ''

    def class_(self, class_):
        if class_ != u"办公":
            return True
        return False

    def handle(self, handle):
        if handle != u"已清除" and handle != u"已删除":
            return True
        return False

    def virus(self, virus):
        if virus == u"蠕虫" or virus == u"勒索" or virus == u"PE":
            return True
        return False

    def exists(self):
        title = [u"告警时间", u"数据录入时间", u"所属分行", u"IP地址", u"MAC地址",u"感染主机名",
                 u"病毒名称", u"病毒类型", u"受感染文件", u"感染源", u"感染路径", u"处理结果", u"感染类型",
                 u"感染机被感染时间", u"扫描方式", u"病毒码组件", u"系统类型"]
        self.date = datetime.datetime.now().strftime("%Y%m%d")
        if os.path.exists("logFile/%s.xlsx" % self.date):
            wb = openpyxl.load_workbook("logFile/%s.xlsx" % self.date)
            try:
                ws = wb.get_sheet_by_name(u"告警日志")
            except:
                wb.close()
                wb = Data_Operat().create_Newsheet("logFile/%s.xlsx" % self.date, [u"告警日志"])
                ws = Data_Operat().get_Sheet(wb, u"告警日志", title)
        else:
            wb = openpyxl.Workbook()
            wb.create_sheet(u"告警日志")
            wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
            ws = wb.get_sheet_by_name(u"告警日志")
            ws.append(title)
        return wb, ws

    def detect(self, data, asset_branch, asset_virus):
        wb, ws = self.exists()
        flag = False
        for d in data:
            equipment, branch, qu, class_ = Data_Operat().system_Belong(ip=d[2], asset=asset_branch)
            virus_ = Data_Operat().virus_Belong(virus=d[5], asset=asset_virus, defaultvirus=u"未知病毒")
            handle_ = Data_Operat().handle_result(code=d[9])
            detect_ = Data_Operat().detect_result(code=d[11])
            if self.class_(class_) or self.handle(handle_) or self.virus(virus_):
                self.clr.print_red_text(u"监测到一个告警,详情如下:")
                self.clr.print_red_text(u"  设  备:     %s-%s-%s" % (branch, qu, equipment))
                self.clr.print_red_text(u"  IP地址:     %s" % d[2])
                self.clr.print_red_text(u"  MAC地址:    %s" % d[3])
                self.clr.print_red_text(u"  主机名:     %s" % d[4])
                self.clr.print_red_text(u"  病毒名:     %s" % d[5])
                self.clr.print_red_text(u"  病毒类型:   %s" % virus_)
                self.clr.print_red_text(u"  感染文件名: %s" % d[6])
                self.clr.print_red_text(u"  感染源:     %s" % d[7])
                self.clr.print_red_text(u"  感染路径:   %s" % d[8])
                self.clr.print_red_text(u"  处理结果:   %s" % handle_)
                self.clr.print_red_text(u"  扫描方式:   %s" % detect_)
                self.clr.print_red_text(u"  病毒码组件: %s" % d[12])
                self.clr.print_red_text(u"  操作系统:   %s" % d[13])
                handle_result = Data_Operat().operat_Belong(handle_)
                flag = True
                time__ = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append(time__, d[0], "%s-%s-%s"%(branch, qu, equipment), d[2], d[3], d[4], d[5], virus_, d[6], d[7],
                          d[8], handle_, handle_result, d[10], detect_, d[12], d[13])
        wb.save("logFile/%s.xlsx" % self.date)
        wb.close()
        return flag