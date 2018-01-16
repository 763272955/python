# -*- coding:utf-8 -*-

import openpyxl
from collections import defaultdict
import os

title=[u"规则名称",u"源IP",u"上报引擎"]
dict_all = {}
rule_index = title.index(u"规则名称")
YIP_index = title.index(u"源IP")
report_index = title.index(u"上报引擎")
dict_all[title[rule_index]] = defaultdict(int)
dict_all[title[YIP_index]] = defaultdict(int)
dict_all[title[report_index]] = defaultdict(int)
# wb = openpyxl.load_workbook(u"E:/工作目录/互联网.xlsx")
listdir = os.listdir(u'C:/Users/user/Desktop/11月份互联网结果/')
# listdir = os.listdir('dir')
wb = openpyxl.Workbook()
wb.create_sheet(u'规则名称')
wb.create_sheet(u"源IP")
wb.create_sheet(u"上报引擎")
ws_rule = wb.get_sheet_by_name(u'规则名称')
ws_YIP = wb.get_sheet_by_name(u"源IP")
ws_report = wb.get_sheet_by_name(u"上报引擎")
# print listdir
for dir in listdir:
    wb_=openpyxl.load_workbook(u'C:/Users/user/Desktop/11月份互联网结果/%s/数据筛选.xlsx' % dir)
    ws = wb_.get_sheet_by_name(u'端口(全)')
for row in list(ws.rows)[1:]:
    dict_all[title[rule_index]][row[1].value] += 1
    dict_all[title[report_index]][row[6].value] += 1
    dict_all[title[YIP_index]][row[2].value] += 1
for key in dict_all[title[rule_index]].keys():
    ws_rule.append([key, dict_all[title[rule_index]][key]])
for key in dict_all[title[YIP_index]].keys():
    ws_YIP.append([key, dict_all[title[YIP_index]][key]])
for key in dict_all[title[report_index]].keys():
    ws_report.append([key, dict_all[title[report_index]][key]])
wb.save(u"综合.xlsx")
wb.close()