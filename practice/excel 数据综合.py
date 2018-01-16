#-*- coding: utf-8 -*-
import re
import openpyxl
import time

dict = {}
title = [u'告警时间', u'规则名称', u'源IP', u'源端口', u'目的IP', u'目的端口',
              u'上报引擎', u'返回消息', u'网口编号', u'网口别名', u'全行']

# ws = wb.get_sheet_by_name('list')
# ws.append(title)
wk = openpyxl.load_workbook(u"outputFile/数据筛选.xlsx")
ws = wk.get_sheet_by_name(u'总和')
ws = list(ws)[1:]
wb = openpyxl.Workbook()
ws1 = wb.create_sheet(u'总和')
ws_end = wb.get_sheet_by_name(u'总和')
for row in ws:
        if row[0].value not in dict.keys():
            dict[row[0].value]=int(row[1].value)
        else:
            dict[row[0].value]+=int(row[1].value)
for key in dict.keys():
    ws1.append([key, dict[key]])
wb.save(u"outputFile/ok.xlsx")
wb.close()




# ws = openpyxl.load_workbook("D:\工作分析\ssq.xlsx")
# for sheet in wb:
#     print sheet.title
# ws = wb.active
# ws1 = wb.create_sheet()
# ws.title = "new title"
# ws.sheet_properties.tabColor = "1072BA"
# ws3 = wb["new title"]
# ws4 = wb.get_sheet_by_name("new title")
# wb.save(u"outputFile/数据筛选.xlsx")
# wb.close()
